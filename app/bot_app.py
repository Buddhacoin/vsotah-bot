import asyncio
import base64
import os
import time
import traceback
from datetime import date, datetime, timedelta
from io import BytesIO

import asyncpg
from aiohttp import web
from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    Message,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
    PreCheckoutQuery,
    LabeledPrice,
    BotCommand,
    LinkPreviewOptions,
    BufferedInputFile,
)

from app.ai.router import (
    ai_router,
    vision_router,
    generate_nano_banana_image,
    generate_gpt_image,
    edit_gpt_image,
    analyze_pdf_images_with_openai,
    file_router,
)

BOT_TOKEN = os.getenv("BOT_TOKEN")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
DATABASE_URL = os.getenv("DATABASE_URL")
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")

OPENAI_TEXT_MODEL = os.getenv("OPENAI_TEXT_MODEL", "gpt-5.4-mini")
OPENAI_VISION_MODEL = os.getenv("OPENAI_VISION_MODEL", "gpt-5.4-mini")
ANTHROPIC_TEXT_MODEL = os.getenv("ANTHROPIC_TEXT_MODEL", "claude-sonnet-4-6")
ANTHROPIC_VISION_MODEL = os.getenv("ANTHROPIC_VISION_MODEL", "claude-sonnet-4-6")
GEMINI_TEXT_MODEL = os.getenv("GEMINI_TEXT_MODEL", "gemini-2.5-flash")
GEMINI_VISION_MODEL = os.getenv("GEMINI_VISION_MODEL", "gemini-2.5-flash")
NANO_BANANA_MODEL = os.getenv("NANO_BANANA_MODEL", "imagen-4.0-generate-001")
GPT_IMAGE_MODEL = os.getenv("GPT_IMAGE_MODEL", "gpt-image-1")
GPT_IMAGE_QUALITY = os.getenv("GPT_IMAGE_QUALITY", "high")

# Speed settings. You can override these in Railway Variables if needed.
TEXT_HISTORY_LIMIT = int(os.getenv("TEXT_HISTORY_LIMIT", "6"))
VISION_HISTORY_LIMIT = int(os.getenv("VISION_HISTORY_LIMIT", "2"))
TEXT_MAX_TOKENS = int(os.getenv("TEXT_MAX_TOKENS", "1200"))
VISION_MAX_TOKENS = int(os.getenv("VISION_MAX_TOKENS", "900"))
AI_TIMEOUT_SECONDS = int(os.getenv("AI_TIMEOUT_SECONDS", "75"))
MAX_DOCUMENT_SIZE = int(os.getenv("MAX_DOCUMENT_SIZE", str(10 * 1024 * 1024)))
FILE_TEXT_LIMIT = int(os.getenv("FILE_TEXT_LIMIT", "18000"))

PORT = int(os.getenv("PORT", "8080"))

ADMIN_IDS = {
    int(x.strip())
    for x in os.getenv("ADMIN_IDS", "").split(",")
    if x.strip().isdigit()
}

FREE_DAILY_LIMIT = 15
FREE_WEEKLY_LIMIT = 105

PLAN_WEEKLY_LIMITS = {
    "FREE": 105,
    "PLUS": 500,
    "PRO": 1400,
    "VIP": None,
}

TARIFFS = {
    "PLUS": {
        "title": "PLUS",
        "description": "500 запросов в неделю",
        "prices": {1: 199, 3: 400, 6: 800, 12: 1600},
    },
    "PRO": {
        "title": "PRO",
        "description": "1400 запросов в неделю",
        "prices": {1: 499, 3: 1000, 6: 2000, 12: 3000},
    },
    "VIP": {
        "title": "VIP",
        "description": "Безлимит",
        "prices": {1: 1499, 3: 3000, 6: 6000, 12: 9900},
    },
}

SPAM_WINDOW_SECONDS = 8
SPAM_MAX_MESSAGES = 5
SPAM_BLOCK_SECONDS = 60

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


db_pool = None
recent_starts = {}


def no_preview():
    return LinkPreviewOptions(is_disabled=True)


def main_menu():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="👤 Профиль", callback_data="profile"),
                InlineKeyboardButton(text="💳 Купить подписку", callback_data="premium"),
            ],
            [InlineKeyboardButton(text="🤖 Выбрать нейросеть", callback_data="models")],
            [InlineKeyboardButton(text="🧠 Наши каналы", callback_data="channels")],
        ]
    )


def models_menu():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🌀 ChatGPT — GPT-4o mini 🟢 FREE", callback_data="set_model_gpt")],
            [InlineKeyboardButton(text="✦ Gemini — 2.5 Flash 🟢 FREE", callback_data="set_model_gemini")],
            [InlineKeyboardButton(text="✴️ Claude — Sonnet ⭐ PLUS", callback_data="set_model_claude")],
            [InlineKeyboardButton(text="🍌 Nano Banana Pro 💎 PRO", callback_data="set_model_nanobanana")],
            [InlineKeyboardButton(text="🌀 Sora GPT Image 👑 VIP", callback_data="set_model_gptimage")],
            [InlineKeyboardButton(text="← Назад", callback_data="back_main")],
        ]
    )


def tariffs_menu():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="⭐ PLUS — 500 запросов / неделя", callback_data="tariff_PLUS")],
            [InlineKeyboardButton(text="💎 PRO — 1400 запросов / неделя", callback_data="tariff_PRO")],
            [InlineKeyboardButton(text="👑 VIP — безлимит", callback_data="tariff_VIP")],
            [InlineKeyboardButton(text="← Назад", callback_data="back_main")],
        ]
    )


def channels_menu():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="⚡ Молния Live", url="https://t.me/MolniyaLiveNews")],
            [InlineKeyboardButton(text="⚡ Молния News", url="https://t.me/LightningNewsSupport")],
            [InlineKeyboardButton(text="← Назад", callback_data="back_main")],
        ]
    )


def period_menu(plan: str):
    prices = TARIFFS[plan]["prices"]
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=f"1 месяц — ⭐ {prices[1]}", callback_data=f"period_{plan}_1")],
            [InlineKeyboardButton(text=f"3 месяца — ⭐ {prices[3]}", callback_data=f"period_{plan}_3")],
            [InlineKeyboardButton(text=f"6 месяцев — ⭐ {prices[6]}", callback_data=f"period_{plan}_6")],
            [InlineKeyboardButton(text=f"12 месяцев — ⭐ {prices[12]}", callback_data=f"period_{plan}_12")],
            [InlineKeyboardButton(text="← Назад", callback_data="premium")],
        ]
    )


def payment_method_menu(plan: str, months: int):
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="💳 Карта / СБП — скоро", callback_data="rub_disabled")],
            [InlineKeyboardButton(text="⭐ Telegram Stars", callback_data=f"pay_stars_{plan}_{months}")],
            [InlineKeyboardButton(text="← Назад", callback_data=f"tariff_{plan}")],
        ]
    )


def get_week_start():
    today = date.today()
    return today - timedelta(days=today.weekday())


def add_months_rough(months: int):
    return datetime.utcnow() + timedelta(days=30 * months)


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


PLAN_LEVELS = {
    "FREE": 0,
    "PLUS": 1,
    "PRO": 2,
    "VIP": 3,
}

MODEL_REQUIRED_PLAN = {
    "gpt": "FREE",
    "gemini": "FREE",
    "claude": "PLUS",
    "nanobanana": "PRO",
    "gptimage": "VIP",
}


def plan_level(plan: str | None) -> int:
    return PLAN_LEVELS.get((plan or "FREE").upper(), 0)


def model_required_plan(model: str) -> str:
    return MODEL_REQUIRED_PLAN.get(model, "FREE")


def has_model_access(user_plan: str | None, model: str) -> bool:
    return plan_level(user_plan) >= plan_level(model_required_plan(model))


def model_display_name(model: str) -> str:
    names = {
        "gpt": "🌀 ChatGPT — GPT-4o mini 🟢 FREE",
        "gemini": "✦ Gemini — 2.5 Flash 🟢 FREE",
        "claude": "✴️ Claude — Sonnet ⭐ PLUS",
        "nanobanana": "🍌 Nano Banana Pro 💎 PRO",
        "gptimage": "🌀 Sora GPT Image 👑 VIP",
    }
    return names.get(model, "🌀 ChatGPT — GPT-4o mini 🟢 FREE")


def premium_required_text(model: str) -> str:
    required = model_required_plan(model)
    return (
        f"🔒 {model_display_name(model)} доступна с тарифа {required}.\n\n"
        "Выберите подходящий тариф, чтобы открыть более мощные нейросети."
    )


def welcome_text():
    return """👋 Добро пожаловать в @GPTclaudeAIbot

Ваш AI-бот для работы с нейросетями в одном месте.

📝 Генерация текста:
• ChatGPT
• Claude
• Gemini

🌇 Генерация изображений:
• Nano Banana Pro
• Sora GPT Image

📷 Анализ фото:
• вопросы по изображениям
• распознавание текста
• помощь с заданиями по фото

🧠 Наши каналы:
• Наш канал: <a href='https://t.me/MolniyaLiveNews'>Молния Live</a>
• Канал support: <a href='https://t.me/LightningNewsSupport'>Молния News</a>

Напишите вопрос, отправьте фото или выберите действие ниже."""


def premium_text():
    return """💳 Купить подписку

🟢 FREE
• ChatGPT — GPT-4o mini
• Gemini — 2.5 Flash
• 15 запросов в день / 105 в неделю

⭐ PLUS — 500 запросов в неделю
• Claude Sonnet
• больше лимитов

💎 PRO — 1400 запросов в неделю
• Nano Banana Pro
• генерация изображений

👑 VIP — безлимит
• Sora GPT Image
• максимум возможностей"""


def channels_text():
    return """🧠 Наши каналы:

• Наш канал: <a href='https://t.me/MolniyaLiveNews'>Молния Live</a>
• Канал support: <a href='https://t.me/LightningNewsSupport'>Молния News</a>"""


async def init_db():
    global db_pool
    db_pool = await asyncpg.create_pool(DATABASE_URL)

    async with db_pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                telegram_id BIGINT PRIMARY KEY,
                username TEXT,
                first_name TEXT,
                plan TEXT DEFAULT 'FREE',
                selected_model TEXT DEFAULT 'gpt',
                daily_used INTEGER DEFAULT 0,
                weekly_used INTEGER DEFAULT 0,
                day_start DATE DEFAULT CURRENT_DATE,
                week_start DATE DEFAULT CURRENT_DATE,
                created_at TIMESTAMP DEFAULT NOW()
            );
        """)
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS plan_until TIMESTAMP;")
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS selected_model TEXT DEFAULT 'gpt';")
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS daily_used INTEGER DEFAULT 0;")
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS weekly_used INTEGER DEFAULT 0;")
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS day_start DATE DEFAULT CURRENT_DATE;")
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS week_start DATE DEFAULT CURRENT_DATE;")

        await conn.execute("""
            CREATE TABLE IF NOT EXISTS messages (
                id SERIAL PRIMARY KEY,
                telegram_id BIGINT NOT NULL,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            );
        """)

        await conn.execute("""
            CREATE TABLE IF NOT EXISTS payments (
                id SERIAL PRIMARY KEY,
                telegram_id BIGINT NOT NULL,
                plan TEXT NOT NULL,
                amount INTEGER NOT NULL,
                currency TEXT NOT NULL,
                payload TEXT NOT NULL,
                telegram_payment_charge_id TEXT,
                provider_payment_charge_id TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            );
        """)
        await conn.execute("ALTER TABLE payments ADD COLUMN IF NOT EXISTS months INTEGER DEFAULT 1;")

        await conn.execute("""
            CREATE TABLE IF NOT EXISTS spam_state (
                telegram_id BIGINT PRIMARY KEY,
                window_start BIGINT DEFAULT 0,
                message_count INTEGER DEFAULT 0,
                blocked_until BIGINT DEFAULT 0
            );
        """)

        await conn.execute("""
            CREATE TABLE IF NOT EXISTS events (
                id SERIAL PRIMARY KEY,
                telegram_id BIGINT,
                event_type TEXT NOT NULL,
                details TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            );
        """)


async def log_event(telegram_id: int | None, event_type: str, details: str = ""):
    try:
        async with db_pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO events (telegram_id, event_type, details) VALUES ($1, $2, $3)",
                telegram_id,
                event_type,
                details[:1000],
            )
    except Exception as e:
        print(f"LOG EVENT ERROR: {e}")


async def setup_bot_info():
    await bot.set_my_commands([
        BotCommand(command="start", description="👋 Что умеет бот"),
        BotCommand(command="account", description="👤 Мой профиль"),
        BotCommand(command="premium", description="💳 Купить подписку"),
        BotCommand(command="models", description="🤖 Выбрать нейросеть"),
        BotCommand(command="channels", description="🧠 Наши каналы"),
        BotCommand(command="deletecontext", description="💬 Удалить контекст"),
    ])


async def get_or_create_user_by_data(telegram_id, username=None, first_name=None):
    today = date.today()
    week_start = get_week_start()

    async with db_pool.acquire() as conn:
        user = await conn.fetchrow("SELECT * FROM users WHERE telegram_id=$1", telegram_id)

        if not user:
            await conn.execute("""
                INSERT INTO users (telegram_id, username, first_name, day_start, week_start)
                VALUES ($1, $2, $3, $4, $5)
            """, telegram_id, username, first_name, today, week_start)
            await log_event(telegram_id, "new_user", username or "")
        else:
            await conn.execute("UPDATE users SET username=$2, first_name=$3 WHERE telegram_id=$1", telegram_id, username, first_name)

        user = await conn.fetchrow("SELECT * FROM users WHERE telegram_id=$1", telegram_id)

        if user["day_start"] != today:
            await conn.execute("UPDATE users SET daily_used=0, day_start=$2 WHERE telegram_id=$1", telegram_id, today)

        if user["week_start"] != week_start:
            await conn.execute("UPDATE users SET weekly_used=0, week_start=$2 WHERE telegram_id=$1", telegram_id, week_start)

        user = await conn.fetchrow("SELECT * FROM users WHERE telegram_id=$1", telegram_id)

        if user["plan"] != "FREE" and user["plan_until"] and user["plan_until"] < datetime.utcnow():
            await conn.execute("UPDATE users SET plan='FREE', plan_until=NULL WHERE telegram_id=$1", telegram_id)
            await log_event(telegram_id, "plan_expired", user["plan"])

        return await conn.fetchrow("SELECT * FROM users WHERE telegram_id=$1", telegram_id)


async def get_or_create_user(message: Message):
    return await get_or_create_user_by_data(
        telegram_id=message.from_user.id,
        username=message.from_user.username,
        first_name=message.from_user.first_name,
    )


async def check_spam(telegram_id: int):
    now = int(time.time())

    async with db_pool.acquire() as conn:
        state = await conn.fetchrow("SELECT * FROM spam_state WHERE telegram_id=$1", telegram_id)

        if not state:
            await conn.execute("""
                INSERT INTO spam_state (telegram_id, window_start, message_count, blocked_until)
                VALUES ($1, $2, 1, 0)
            """, telegram_id, now)
            return True, None

        if state["blocked_until"] and state["blocked_until"] > now:
            return False, state["blocked_until"] - now

        if now - state["window_start"] > SPAM_WINDOW_SECONDS:
            await conn.execute("UPDATE spam_state SET window_start=$2, message_count=1, blocked_until=0 WHERE telegram_id=$1", telegram_id, now)
            return True, None

        new_count = state["message_count"] + 1

        if new_count > SPAM_MAX_MESSAGES:
            blocked_until = now + SPAM_BLOCK_SECONDS
            await conn.execute("UPDATE spam_state SET message_count=$2, blocked_until=$3 WHERE telegram_id=$1", telegram_id, new_count, blocked_until)
            await log_event(telegram_id, "spam_block", str(SPAM_BLOCK_SECONDS))
            return False, SPAM_BLOCK_SECONDS

        await conn.execute("UPDATE spam_state SET message_count=$2 WHERE telegram_id=$1", telegram_id, new_count)

    return True, None


async def check_limit(user):
    plan = user["plan"]

    if plan == "VIP":
        return True, None

    if plan == "FREE":
        if user["daily_used"] >= FREE_DAILY_LIMIT:
            return False, "FREE_DAY_LIMIT"
        if user["weekly_used"] >= FREE_WEEKLY_LIMIT:
            return False, "FREE_WEEK_LIMIT"
        return True, None

    weekly_limit = PLAN_WEEKLY_LIMITS.get(plan)
    if weekly_limit is not None and user["weekly_used"] >= weekly_limit:
        return False, f"{plan}_WEEK_LIMIT"

    return True, None


async def increase_usage(telegram_id):
    async with db_pool.acquire() as conn:
        await conn.execute("""
            UPDATE users
            SET daily_used = daily_used + 1,
                weekly_used = weekly_used + 1
            WHERE telegram_id=$1
        """, telegram_id)


async def save_message(telegram_id, role, content):
    async with db_pool.acquire() as conn:
        await conn.execute(
            "INSERT INTO messages (telegram_id, role, content) VALUES ($1, $2, $3)",
            telegram_id,
            role,
            content[:12000],
        )


async def get_chat_history(telegram_id, limit=TEXT_HISTORY_LIMIT):
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT role, content
            FROM messages
            WHERE telegram_id=$1
            ORDER BY created_at DESC
            LIMIT $2
        """, telegram_id, limit)

    history = []
    for row in reversed(rows):
        if row["role"] in {"user", "assistant", "system"}:
            history.append({"role": row["role"], "content": row["content"]})
    return history


async def clear_chat(telegram_id):
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM messages WHERE telegram_id=$1", telegram_id)
    await log_event(telegram_id, "delete_context")


async def activate_plan(telegram_id: int, plan: str, months: int):
    until = add_months_rough(months)
    async with db_pool.acquire() as conn:
        await conn.execute("""
            UPDATE users
            SET plan=$2, plan_until=$3, daily_used=0, weekly_used=0
            WHERE telegram_id=$1
        """, telegram_id, plan, until)
    await log_event(telegram_id, "activate_plan", f"{plan} {months} months")


async def user_profile_text(user):
    model_names = {
        "gpt": "ChatGPT",
        "claude": "Claude",
        "gemini": "Gemini",
        "nanobanana": "Nano Banana",
        "gptimage": "Sora GPT Image",
        "deepseek": "DeepSeek",
    }

    plan = user["plan"] or "FREE"
    current_model = model_display_name(user["selected_model"])

    if plan == "VIP":
        usage_block = "Запросов: ♾ безлимит"
    elif plan in {"PLUS", "PRO"}:
        usage_block = f"Запросов в неделю: {user['weekly_used']}/{PLAN_WEEKLY_LIMITS[plan]}"
    else:
        usage_block = (
            f"Запросов сегодня: {user['daily_used']}/{FREE_DAILY_LIMIT}\n"
            f"Запросов в неделю: {user['weekly_used']}/{FREE_WEEKLY_LIMIT}"
        )

    text = (
        f"📊 Статистика использования\n\n"
        f"{usage_block}\n\n"
        f"Подписка: {plan}\n"
        f"Выбрана модель: {current_model}\n\n"
    )

    if plan != "FREE" and user["plan_until"]:
        text += f"Активна до: {user['plan_until'].strftime('%d.%m.%Y')}\n\n"

    text += (
        "Нужно больше? 🚀 Выберите тариф для покупки Premium:\n\n"
        "⭐ PLUS — Claude Sonnet и 500 запросов в неделю\n"
        "💎 PRO — Nano Banana Pro и 1400 запросов в неделю\n"
        "👑 VIP — Sora GPT Image и безлимит"
    )
    return text












async def download_telegram_photo(message: Message) -> bytes:
    photo = message.photo[-1]
    file = await bot.get_file(photo.file_id)
    buffer = BytesIO()
    await bot.download_file(file.file_path, destination=buffer)
    return buffer.getvalue()














async def download_telegram_document(message: Message) -> tuple[str, bytes]:
    document = message.document
    filename = document.file_name or "file"

    if document.file_size and document.file_size > MAX_DOCUMENT_SIZE:
        raise ValueError("FILE_TOO_LARGE")

    file = await bot.get_file(document.file_id)
    buffer = BytesIO()
    await bot.download_file(file.file_path, destination=buffer)
    return filename, buffer.getvalue()


def extract_document_text(filename: str, file_bytes: bytes) -> str:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext in {"txt", "md", "csv", "log"}:
        for encoding in ("utf-8", "utf-8-sig", "cp1251", "latin-1"):
            try:
                return file_bytes.decode(encoding, errors="replace")[:FILE_TEXT_LIMIT]
            except Exception:
                continue
        return file_bytes.decode("utf-8", errors="replace")[:FILE_TEXT_LIMIT]

    if ext == "pdf":
        pages: list[str] = []

        # 1) pypdf — быстрый вариант для PDF с текстовым слоем
        try:
            from pypdf import PdfReader
            reader = PdfReader(BytesIO(file_bytes))
            for page in reader.pages[:25]:
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(text.strip())
                if len("\n".join(pages)) >= FILE_TEXT_LIMIT:
                    break
        except Exception as e:
            print(f"PDF PYPDF READ ERROR: {e}")

        # 2) pdfplumber — иногда лучше читает таблицы/чеки
        if not "\n".join(pages).strip():
            try:
                import pdfplumber
                with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                    for page in pdf.pages[:25]:
                        text = page.extract_text() or ""
                        if text.strip():
                            pages.append(text.strip())
                        if len("\n".join(pages)) >= FILE_TEXT_LIMIT:
                            break
            except Exception as e:
                print(f"PDF PDFPLUMBER READ ERROR: {e}")

        # 3) PyMuPDF — ещё один fallback для текстового слоя
        if not "\n".join(pages).strip():
            try:
                import fitz
                doc = fitz.open(stream=file_bytes, filetype="pdf")
                for page in doc[:25]:
                    text = page.get_text("text") or ""
                    if text.strip():
                        pages.append(text.strip())
                    if len("\n".join(pages)) >= FILE_TEXT_LIMIT:
                        break
                doc.close()
            except Exception as e:
                print(f"PDF FITZ TEXT READ ERROR: {e}")

        return "\n\n".join(pages)[:FILE_TEXT_LIMIT]

    if ext == "docx":
        try:
            from docx import Document
            doc = Document(BytesIO(file_bytes))
            parts = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(parts)[:FILE_TEXT_LIMIT]
        except Exception as e:
            raise ValueError(f"DOCX_READ_ERROR: {e}")

    if ext in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
            lines = []
            for ws in wb.worksheets[:5]:
                lines.append(f"Лист: {ws.title}")
                for row in ws.iter_rows(max_row=80, values_only=True):
                    values = [str(v) if v is not None else "" for v in row]
                    if any(v.strip() for v in values):
                        lines.append(" | ".join(values))
                    if len("\n".join(lines)) >= FILE_TEXT_LIMIT:
                        return "\n".join(lines)[:FILE_TEXT_LIMIT]
            return "\n".join(lines)[:FILE_TEXT_LIMIT]
        except Exception as e:
            raise ValueError(f"XLSX_READ_ERROR: {e}")

    raise ValueError("UNSUPPORTED_FILE_TYPE")










def short_error_text(error: Exception) -> str:
    return str(error).replace("\n", " ")[:1200]


async def send_ai_error_to_admin(error_text: str):
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, error_text[:3000])
        except Exception:
            pass


async def safe_edit_or_send(callback: CallbackQuery, text: str, reply_markup=None, parse_mode=None):
    try:
        await callback.message.edit_text(
            text,
            reply_markup=reply_markup,
            parse_mode=parse_mode,
            link_preview_options=no_preview(),
        )
    except Exception:
        await callback.message.answer(
            text,
            reply_markup=reply_markup,
            parse_mode=parse_mode,
            link_preview_options=no_preview(),
        )


@dp.message(CommandStart())
async def start_handler(message: Message):
    now = time.time()
    last = recent_starts.get(message.from_user.id, 0)
    if now - last < 2:
        return
    recent_starts[message.from_user.id] = now

    user = await get_or_create_user(message)
    await log_event(message.from_user.id, "start")
    await message.answer(
        f"{welcome_text()}\n\nТекущая нейросеть: {model_display_name(user['selected_model'])}",
        reply_markup=main_menu(),
        parse_mode="HTML",
        link_preview_options=no_preview(),
    )


@dp.message(Command("account"))
async def account_command(message: Message):
    user = await get_or_create_user(message)
    await log_event(message.from_user.id, "account")
    await message.answer(await user_profile_text(user), reply_markup=main_menu())


@dp.message(Command("premium"))
async def premium_command(message: Message):
    await log_event(message.from_user.id, "premium_open")
    await message.answer(premium_text(), reply_markup=tariffs_menu())


@dp.message(Command("models"))
async def models_command(message: Message):
    await log_event(message.from_user.id, "models_command")
    await message.answer(
        "🤖 Выберите нейросеть:\n\n"
        "🟢 FREE доступно всем\n"
        "⭐ PLUS / 💎 PRO / 👑 VIP открываются после подписки",
        reply_markup=models_menu(),
    )


@dp.message(Command("channels"))
async def channels_command(message: Message):
    await log_event(message.from_user.id, "channels_command")
    await message.answer(channels_text(), reply_markup=channels_menu(), parse_mode="HTML", link_preview_options=no_preview())


@dp.message(Command("deletecontext"))
async def delete_context_command(message: Message):
    await clear_chat(message.from_user.id)
    await message.answer("💬 Контекст очищен. Новый чат начат.", reply_markup=main_menu())


@dp.callback_query(F.data == "back_main")
async def back_main_callback(callback: CallbackQuery):
    await callback.answer()
    await safe_edit_or_send(callback, welcome_text(), reply_markup=main_menu(), parse_mode="HTML")


@dp.callback_query(F.data == "profile")
async def profile_callback(callback: CallbackQuery):
    await callback.answer("Открываю профиль...")
    user = await get_or_create_user_by_data(callback.from_user.id, callback.from_user.username, callback.from_user.first_name)
    await log_event(callback.from_user.id, "profile_click")
    await callback.message.answer(await user_profile_text(user), reply_markup=main_menu())


@dp.callback_query(F.data == "channels")
async def channels_callback(callback: CallbackQuery):
    await callback.answer()
    await log_event(callback.from_user.id, "channels_open")
    await safe_edit_or_send(callback, channels_text(), reply_markup=channels_menu(), parse_mode="HTML")


@dp.callback_query(F.data.in_({"premium", "plans"}))
async def premium_callback(callback: CallbackQuery):
    await callback.answer()
    await log_event(callback.from_user.id, "premium_click")
    await safe_edit_or_send(callback, premium_text(), reply_markup=tariffs_menu())


@dp.callback_query(F.data == "models")
async def models_callback(callback: CallbackQuery):
    await callback.answer()
    await log_event(callback.from_user.id, "models_open")
    await safe_edit_or_send(
        callback,
        "🤖 Выберите нейросеть:\n\n"
        "🟢 FREE доступно всем\n"
        "⭐ PLUS / 💎 PRO / 👑 VIP открываются после подписки",
        reply_markup=models_menu(),
    )


@dp.callback_query(F.data.startswith("tariff_"))
async def tariff_callback(callback: CallbackQuery):
    await callback.answer()
    plan = callback.data.replace("tariff_", "")
    if plan not in TARIFFS:
        return
    await log_event(callback.from_user.id, "tariff_select", plan)
    tariff = TARIFFS[plan]
    await safe_edit_or_send(
        callback,
        f"🚀 {tariff['title']}\n\n{tariff['description']}\n\nВыберите период подписки:",
        reply_markup=period_menu(plan),
    )


@dp.callback_query(F.data.startswith("period_"))
async def period_callback(callback: CallbackQuery):
    await callback.answer()
    _, plan, months = callback.data.split("_")
    months = int(months)
    if plan not in TARIFFS:
        return
    price = TARIFFS[plan]["prices"][months]
    await log_event(callback.from_user.id, "period_select", f"{plan} {months}")
    await safe_edit_or_send(
        callback,
        f"💳 Выберите способ оплаты:\n\nТариф: {plan}\nПериод: {months} мес.\nЦена в Stars: ⭐ {price}",
        reply_markup=payment_method_menu(plan, months),
    )


@dp.callback_query(F.data == "rub_disabled")
async def rub_disabled_callback(callback: CallbackQuery):
    await callback.answer("Оплата рублями скоро будет доступна", show_alert=True)


@dp.callback_query(F.data.startswith("pay_stars_"))
async def pay_stars_callback(callback: CallbackQuery):
    await callback.answer()
    _, _, plan, months = callback.data.split("_")
    months = int(months)
    if plan not in TARIFFS:
        return
    price = TARIFFS[plan]["prices"][months]
    payload = f"plan:{plan}:months:{months}:user:{callback.from_user.id}:ts:{int(time.time())}"
    await log_event(callback.from_user.id, "invoice_open", f"{plan} {months} {price}")
    await bot.send_invoice(
        chat_id=callback.message.chat.id,
        title=f"{plan} на {months} мес.",
        description=f"{TARIFFS[plan]['description']}. Подписка на {months} мес.",
        payload=payload,
        provider_token="",
        currency="XTR",
        prices=[LabeledPrice(label=f"{plan} на {months} мес.", amount=price)],
    )


@dp.pre_checkout_query()
async def pre_checkout_handler(pre_checkout_query: PreCheckoutQuery):
    payload = pre_checkout_query.invoice_payload
    if not payload.startswith("plan:"):
        await pre_checkout_query.answer(ok=False, error_message="Некорректный платёж.")
        return
    await pre_checkout_query.answer(ok=True)


@dp.message(F.successful_payment)
async def successful_payment_handler(message: Message):
    payment = message.successful_payment
    payload = payment.invoice_payload
    parts = payload.split(":")
    plan = None
    months = 1
    try:
        plan = parts[1]
        months = int(parts[3])
    except Exception:
        pass

    if plan not in TARIFFS:
        await message.answer("⚠️ Платёж получен, но тариф не распознан. Напишите администратору.")
        return

    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO payments (
                telegram_id, plan, months, amount, currency, payload,
                telegram_payment_charge_id, provider_payment_charge_id
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
        """,
            message.from_user.id,
            plan,
            months,
            payment.total_amount,
            payment.currency,
            payload,
            payment.telegram_payment_charge_id,
            payment.provider_payment_charge_id,
        )

    await activate_plan(message.from_user.id, plan, months)
    await message.answer(f"✅ Оплата прошла успешно!\n\nТариф {plan} активирован на {months} мес.", reply_markup=main_menu())


@dp.callback_query(F.data.startswith("set_model_"))
async def set_model_callback(callback: CallbackQuery):
    await callback.answer()
    model = callback.data.replace("set_model_", "")
    allowed_models = {"gpt", "claude", "gemini", "nanobanana", "gptimage"}

    if model not in allowed_models:
        await safe_edit_or_send(callback, "⚠️ Эта модель сейчас недоступна.", reply_markup=models_menu())
        return

    user = await get_or_create_user_by_data(callback.from_user.id, callback.from_user.username, callback.from_user.first_name)
    if not has_model_access(user["plan"], model):
        await safe_edit_or_send(callback, premium_required_text(model), reply_markup=tariffs_menu())
        return

    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE users SET selected_model=$1 WHERE telegram_id=$2", model, callback.from_user.id)

    await log_event(callback.from_user.id, "model_select", model)

    try:
        await callback.message.delete()
    except Exception:
        pass

    await bot.send_message(
        chat_id=callback.message.chat.id,
        text=f"✅ Нейросеть выбрана:\n\n{model_display_name(model)}\n\nНапишите запрос, отправьте фото или выберите действие ниже.",
        reply_markup=main_menu(),
    )


@dp.message(Command("admin"))
async def admin_handler(message: Message):
    if not is_admin(message.from_user.id):
        return
    await message.answer(
        "🛠 Админ-панель\n\n"
        "/stats — общая статистика\n"
        "/users — последние пользователи\n"
        "/payments — платежи\n"
        "/setplus telegram_id\n"
        "/setpro telegram_id\n"
        "/setvip telegram_id\n"
        "/setfree telegram_id"
    )


@dp.message(Command("stats"))
async def stats_handler(message: Message):
    if not is_admin(message.from_user.id):
        return

    async with db_pool.acquire() as conn:
        total_users = await conn.fetchval("SELECT COUNT(*) FROM users")
        today_users = await conn.fetchval("SELECT COUNT(*) FROM users WHERE created_at::date = CURRENT_DATE")
        total_messages = await conn.fetchval("SELECT COUNT(*) FROM messages WHERE role='user'")
        today_messages = await conn.fetchval("SELECT COUNT(*) FROM messages WHERE role='user' AND created_at::date = CURRENT_DATE")
        plus_users = await conn.fetchval("SELECT COUNT(*) FROM users WHERE plan='PLUS'")
        pro_users = await conn.fetchval("SELECT COUNT(*) FROM users WHERE plan='PRO'")
        vip_users = await conn.fetchval("SELECT COUNT(*) FROM users WHERE plan='VIP'")
        total_stars = await conn.fetchval("SELECT COALESCE(SUM(amount), 0) FROM payments")
        starts_today = await conn.fetchval("SELECT COUNT(*) FROM events WHERE event_type='start' AND created_at::date = CURRENT_DATE")
        premium_clicks = await conn.fetchval("SELECT COUNT(*) FROM events WHERE event_type='premium_click'")
        invoices = await conn.fetchval("SELECT COUNT(*) FROM events WHERE event_type='invoice_open'")
        vision_requests = await conn.fetchval("SELECT COUNT(*) FROM events WHERE event_type='ai_vision'")

    await message.answer(
        "📊 Статистика бота\n\n"
        f"Пользователей всего: {total_users}\n"
        f"Новых сегодня: {today_users}\n"
        f"Стартов сегодня: {starts_today}\n\n"
        f"Сообщений всего: {total_messages}\n"
        f"Сообщений сегодня: {today_messages}\n"
        f"Фото-запросов всего: {vision_requests}\n\n"
        f"PLUS: {plus_users}\n"
        f"PRO: {pro_users}\n"
        f"VIP: {vip_users}\n\n"
        f"Открытий премиума: {premium_clicks}\n"
        f"Открытий оплаты Stars: {invoices}\n"
        f"Stars получено: {total_stars}"
    )


@dp.message(Command("users"))
async def users_handler(message: Message):
    if not is_admin(message.from_user.id):
        return

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT telegram_id, username, first_name, plan, weekly_used, created_at
            FROM users ORDER BY created_at DESC LIMIT 15
        """)

    text = "👥 Последние пользователи\n\n"
    for row in rows:
        name = row["username"] or row["first_name"] or "без имени"
        text += (
            f"ID: {row['telegram_id']}\n"
            f"Имя: {name}\n"
            f"Тариф: {row['plan']}\n"
            f"Запросов за неделю: {row['weekly_used']}\n"
            f"Дата: {row['created_at'].strftime('%d.%m %H:%M')}\n\n"
        )
    await message.answer(text[:3900])


@dp.message(Command("payments"))
async def payments_handler(message: Message):
    if not is_admin(message.from_user.id):
        return

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT telegram_id, plan, months, amount, currency, created_at
            FROM payments ORDER BY created_at DESC LIMIT 15
        """)

    if not rows:
        await message.answer("Платежей пока нет.")
        return

    text = "💳 Последние платежи Telegram Stars\n\n"
    for row in rows:
        text += (
            f"ID: {row['telegram_id']}\n"
            f"Тариф: {row['plan']} на {row['months']} мес.\n"
            f"Сумма: {row['amount']} {row['currency']}\n"
            f"Дата: {row['created_at'].strftime('%d.%m %H:%M')}\n\n"
        )
    await message.answer(text[:3900])


async def admin_set_plan(message: Message, plan: str):
    if not is_admin(message.from_user.id):
        return

    parts = message.text.split()
    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Формат команды: /setpro telegram_id")
        return

    telegram_id = int(parts[1])
    until = add_months_rough(1) if plan != "FREE" else None

    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE users SET plan=$2, plan_until=$3 WHERE telegram_id=$1", telegram_id, plan, until)

    await log_event(telegram_id, "admin_set_plan", plan)
    await message.answer(f"✅ Пользователю {telegram_id} установлен тариф {plan}.")


@dp.message(Command("setplus"))
async def setplus_handler(message: Message):
    await admin_set_plan(message, "PLUS")


@dp.message(Command("setpro"))
async def setpro_handler(message: Message):
    await admin_set_plan(message, "PRO")


@dp.message(Command("setvip"))
async def setvip_handler(message: Message):
    await admin_set_plan(message, "VIP")


@dp.message(Command("setfree"))
async def setfree_handler(message: Message):
    await admin_set_plan(message, "FREE")


@dp.message(F.photo)
async def photo_handler(message: Message):
    user = await get_or_create_user(message)
    spam_allowed, wait_seconds = await check_spam(message.from_user.id)

    if not spam_allowed:
        await message.answer(f"🛡 Слишком много сообщений подряд.\n\nПопробуйте снова через {wait_seconds} сек.")
        return

    allowed, reason = await check_limit(user)
    if not allowed:
        await log_event(message.from_user.id, "limit_reached", reason)
        await message.answer("⏳ Лимит сообщений закончился.\n\nВы можете перейти на PLUS, PRO или VIP.", reply_markup=tariffs_menu())
        return

    selected_model = user["selected_model"]
    if not has_model_access(user["plan"], selected_model):
        selected_model = "gpt"
        async with db_pool.acquire() as conn:
            await conn.execute("UPDATE users SET selected_model='gpt' WHERE telegram_id=$1", message.from_user.id)
        await message.answer(
            "ℹ️ Ваш тариф изменился, поэтому я переключил нейросеть на ChatGPT — GPT-4o mini 🟢 FREE."
        )

    is_image_edit = selected_model == "gptimage"
    wait_message = await message.answer("🖼 Редактирую изображение..." if is_image_edit else "📷 Анализирую фото...")

    try:
        question = message.caption or ("Улучши это изображение и сохрани смысл." if is_image_edit else "Что изображено на фото?")
        image_bytes = await download_telegram_photo(message)

        if is_image_edit:
            await save_message(message.from_user.id, "user", f"[Редактирование изображения] {question}")
            edited_bytes, text_note = await edit_gpt_image(question, image_bytes)

            if not edited_bytes:
                await wait_message.edit_text(
                    text_note or "⚠️ Редактирование не вернуло изображение. Попробуйте другой запрос.",
                    reply_markup=main_menu(),
                )
                return

            photo = BufferedInputFile(edited_bytes, filename="edited_gpt_image.png")
            await wait_message.delete()
            await message.answer_photo(photo=photo, caption=(text_note[:900] if text_note else "✅ Готово"))
            await save_message(message.from_user.id, "assistant", "[gptimage image edited]")
            await increase_usage(message.from_user.id)
            await log_event(message.from_user.id, "ai_image_edit", selected_model)
            return

        await save_message(message.from_user.id, "user", f"[Фото] {question}")
        history = await get_chat_history(message.from_user.id)
        answer = await vision_router(selected_model, question, image_bytes, history)

        if not answer:
            answer = "⚠️ AI не смог проанализировать фото. Попробуйте отправить другое изображение или добавьте вопрос текстом."

        await save_message(message.from_user.id, "assistant", answer)
        await increase_usage(message.from_user.id)
        await log_event(message.from_user.id, "ai_vision", selected_model)

        if len(answer) <= 3900:
            await wait_message.edit_text(answer)
        else:
            await wait_message.edit_text(answer[:3900])
            for i in range(3900, len(answer), 3900):
                await message.answer(answer[i:i + 3900])

    except Exception as e:
        admin_error = short_error_text(e)
        print(f"VISION ERROR SHORT:\n{admin_error}")
        print(f"VISION ERROR TRACE:\n{traceback.format_exc()}")

        try:
            await wait_message.edit_text(
                "⚠️ Не удалось обработать фото.\n\n"
                "Попробуйте ещё раз или выберите другую нейросеть.",
                reply_markup=main_menu(),
            )
        except Exception:
            await message.answer(
                "⚠️ Не удалось обработать фото.\n\n"
                "Попробуйте ещё раз или выберите другую нейросеть.",
                reply_markup=main_menu(),
            )


@dp.message(F.document)
async def document_handler(message: Message):
    user = await get_or_create_user(message)
    spam_allowed, wait_seconds = await check_spam(message.from_user.id)

    if not spam_allowed:
        await message.answer(f"🛡 Слишком много сообщений подряд.\n\nПопробуйте снова через {wait_seconds} сек.")
        return

    allowed, reason = await check_limit(user)
    if not allowed:
        await log_event(message.from_user.id, "limit_reached", reason)
        await message.answer("⏳ Лимит сообщений закончился.\n\nВы можете перейти на PLUS, PRO или VIP.", reply_markup=tariffs_menu())
        return

    selected_model = user["selected_model"]
    if selected_model in {"nanobanana", "gptimage"}:
        await message.answer(
            "📎 Для анализа файлов выберите ChatGPT, Claude или Gemini в меню «Выбрать нейросеть».",
            reply_markup=main_menu(),
        )
        return

    if not has_model_access(user["plan"], selected_model):
        selected_model = "gpt"
        async with db_pool.acquire() as conn:
            await conn.execute("UPDATE users SET selected_model='gpt' WHERE telegram_id=$1", message.from_user.id)
        await message.answer(
            "ℹ️ Ваш тариф изменился, поэтому я переключил нейросеть на ChatGPT — GPT-4o mini 🟢 FREE."
        )

    wait_message = await message.answer("📎 Читаю файл...")

    try:
        filename, file_bytes = await download_telegram_document(message)
        question = message.caption or "Проанализируй файл и выдели главное."

        try:
            extracted_text = await asyncio.to_thread(extract_document_text, filename, file_bytes)
        except ValueError as e:
            error_code = str(e)
            if error_code == "FILE_TOO_LARGE":
                await wait_message.edit_text("⚠️ Файл слишком большой. Сейчас лимит — до 10 МБ.", reply_markup=main_menu())
            elif error_code == "UNSUPPORTED_FILE_TYPE":
                await wait_message.edit_text("⚠️ Пока поддерживаются TXT, PDF, DOCX и XLSX.", reply_markup=main_menu())
            else:
                await wait_message.edit_text(
                    "⚠️ Не удалось прочитать файл. Попробуйте другой файл или отправьте текстом.",
                    reply_markup=main_menu(),
                )
            return

        await save_message(message.from_user.id, "user", f"[Файл {filename}] {question}")
        history = await get_chat_history(message.from_user.id)

        if not extracted_text.strip():
            if filename.lower().endswith(".pdf"):
                await wait_message.edit_text("📄 PDF похож на скан. Анализирую страницы как изображения...")
                answer = await analyze_pdf_images_with_openai(question, filename, file_bytes, history)
                if not answer:
                    await wait_message.edit_text(
                        "⚠️ Не удалось извлечь текст из PDF. Возможно, файл защищён или страницы плохо читаются.",
                        reply_markup=main_menu(),
                    )
                    return
            else:
                await wait_message.edit_text(
                    "⚠️ Не удалось извлечь текст из файла. Возможно, это скан или защищённый документ.",
                    reply_markup=main_menu(),
                )
                return
        else:
            await wait_message.edit_text("🧠 Анализирую файл...")
            answer = await file_router(selected_model, question, filename, extracted_text, history)

        if not answer:
            answer = "⚠️ AI вернул пустой ответ. Попробуйте задать вопрос по файлу точнее."

        await save_message(message.from_user.id, "assistant", answer)
        await increase_usage(message.from_user.id)
        await log_event(message.from_user.id, "ai_file", selected_model)

        if len(answer) <= 3900:
            await wait_message.edit_text(answer)
        else:
            await wait_message.edit_text(answer[:3900])
            for i in range(3900, len(answer), 3900):
                await message.answer(answer[i:i + 3900])

    except Exception as e:
        admin_error = short_error_text(e)
        print(f"FILE ERROR SHORT:\n{admin_error}")
        print(f"FILE ERROR TRACE:\n{traceback.format_exc()}")

        try:
            await wait_message.edit_text(
                "⚠️ Не удалось обработать файл. Попробуйте ещё раз или отправьте файл в другом формате.",
                reply_markup=main_menu(),
            )
        except Exception:
            await message.answer(
                "⚠️ Не удалось обработать файл. Попробуйте ещё раз или отправьте файл в другом формате.",
                reply_markup=main_menu(),
            )


@dp.message(F.text)
async def chat_handler(message: Message):
    if message.text.startswith("/"):
        return

    user = await get_or_create_user(message)
    spam_allowed, wait_seconds = await check_spam(message.from_user.id)

    if not spam_allowed:
        await message.answer(f"🛡 Слишком много сообщений подряд.\n\nПопробуйте снова через {wait_seconds} сек.")
        return

    allowed, reason = await check_limit(user)
    if not allowed:
        await log_event(message.from_user.id, "limit_reached", reason)
        await message.answer("⏳ Лимит сообщений закончился.\n\nВы можете перейти на PLUS, PRO или VIP.", reply_markup=tariffs_menu())
        return

    selected_model = user["selected_model"]
    if not has_model_access(user["plan"], selected_model):
        selected_model = "gpt"
        async with db_pool.acquire() as conn:
            await conn.execute("UPDATE users SET selected_model='gpt' WHERE telegram_id=$1", message.from_user.id)
        await message.answer(
            "ℹ️ Ваш тариф изменился, поэтому я переключил нейросеть на ChatGPT — GPT-4o mini 🟢 FREE."
        )

    if selected_model in {"nanobanana", "gptimage"}:
        wait_text = "🍌 Генерирую изображение..." if selected_model == "nanobanana" else "🌀 Генерирую изображение..."
        wait_message = await message.answer(wait_text)
        try:
            await save_message(message.from_user.id, "user", message.text)

            if selected_model == "nanobanana":
                image_bytes, text_note = await generate_nano_banana_image(message.text)
            else:
                image_bytes, text_note = await generate_gpt_image(message.text)

            if not image_bytes:
                await wait_message.edit_text(
                    text_note or "⚠️ Генерация не вернула изображение. Попробуйте другой запрос.",
                    reply_markup=main_menu(),
                )
                return

            filename = "nano_banana.png" if selected_model == "nanobanana" else "sora_gpt_image.png"
            photo = BufferedInputFile(image_bytes, filename=filename)
            await wait_message.delete()
            await message.answer_photo(
                photo=photo,
                caption=(text_note[:900] if text_note else "✅ Готово"),
            )

            await save_message(message.from_user.id, "assistant", f"[{selected_model} image generated]")
            await increase_usage(message.from_user.id)
            await log_event(message.from_user.id, "ai_image", selected_model)
            return

        except Exception as e:
            admin_error = short_error_text(e)
            print(f"IMAGE ERROR SHORT:\n{admin_error}")
            print(f"IMAGE ERROR TRACE:\n{traceback.format_exc()}")
            await send_ai_error_to_admin(f"⚠️ AI ERROR | {selected_model} | {admin_error}")
            try:
                await wait_message.edit_text(
                    "⚠️ Генерация изображений временно недоступна.\n\n"
                    "Попробуйте позже или выберите другую нейросеть.",
                    reply_markup=main_menu(),
                )
            except Exception:
                await message.answer(
                    "⚠️ Генерация изображений временно недоступна.\n\n"
                    "Попробуйте позже или выберите другую нейросеть.",
                    reply_markup=main_menu(),
                )
            return

    wait_message = await message.answer("Печатает ответ...")

    try:
        await save_message(message.from_user.id, "user", message.text)
        history = await get_chat_history(message.from_user.id)
        answer = await ai_router(selected_model, history)

        if not answer:
            answer = "⚠️ AI вернул пустой ответ. Попробуйте переформулировать вопрос."

        await save_message(message.from_user.id, "assistant", answer)
        await increase_usage(message.from_user.id)
        await log_event(message.from_user.id, "ai_message", selected_model)

        if len(answer) <= 3900:
            await wait_message.edit_text(answer)
        else:
            await wait_message.edit_text(answer[:3900])
            for i in range(3900, len(answer), 3900):
                await message.answer(answer[i:i + 3900])

    except Exception as e:
        admin_error = short_error_text(e)
        print(f"AI ERROR SHORT:\n{admin_error}")
        print(f"AI ERROR TRACE:\n{traceback.format_exc()}")

        await send_ai_error_to_admin(f"⚠️ AI ERROR | {selected_model} | {admin_error}")

        try:
            await wait_message.edit_text(
                "⚠️ Сейчас выбранная нейросеть временно недоступна.\n\n"
                "Попробуйте позже или выберите другую нейросеть.",
                reply_markup=main_menu(),
            )
        except Exception:
            await message.answer(
                "⚠️ Сейчас выбранная нейросеть временно недоступна.\n\n"
                "Попробуйте позже или выберите другую нейросеть.",
                reply_markup=main_menu(),
            )


async def health_handler(request):
    return web.json_response({"ok": True, "service": "gptclaude-bot"})


async def tribute_webhook_handler(request):
    try:
        data = await request.json()
        print(f"TRIBUTE WEBHOOK RECEIVED: {str(data)[:1000]}")
        await log_event(None, "tribute_webhook_received", str(data)[:1000])
    except Exception as e:
        print(f"TRIBUTE WEBHOOK ERROR: {e}")
    return web.json_response({"ok": True})


async def start_web_server():
    app = web.Application()
    app.router.add_get("/", health_handler)
    app.router.add_get("/health", health_handler)
    app.router.add_post("/tribute-webhook", tribute_webhook_handler)

    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    print(f"WEB SERVER STARTED ON PORT {PORT}")


async def main():
    print("BOT STARTING")

    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN is missing")
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY is missing")
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL is missing")

    await bot.delete_webhook(drop_pending_updates=True)
    await init_db()
    await setup_bot_info()
    await start_web_server()

    print("DATABASE CONNECTED")
    print("BOT STARTED")

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())

